#!/usr/bin/env python3
"""Build merged TC xlsx from tc_rows.json + optional 001 xlsx."""
import json
import shutil
import sys
from copy import copy
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[4]
TEMPLATE = ROOT / ".agents/skills/tc-example-generator/templates/self-test-template.xlsx"
ROWS_JSON = Path(__file__).parent / "tc_rows_002_010.json"
XLSX_001 = ROOT / "specs/001-m365-email-send/TC_001_m365-email-send_20260610.xlsx"
OUT = ROOT / f"TC_001-010_merged_{date.today().strftime('%Y%m%d')}.xlsx"


def load_001_rows():
    if not XLSX_001.exists():
        return []
    wb = load_workbook(XLSX_001, read_only=True)
    ws = wb["Account API"]
    rows = []
    for r in range(2, ws.max_row + 1):
        d = ws.cell(r, 4).value
        if not d:
            continue
        rows.append(
            {
                "spec": "001",
                "D": d,
                "E": ws.cell(r, 5).value,
                "F": ws.cell(r, 6).value,
                "G": ws.cell(r, 7).value,
                "H": ws.cell(r, 8).value,
            }
        )
    wb.close()
    return rows


def write_xlsx(rows):
    shutil.copy(TEMPLATE, OUT)
    wb = load_workbook(OUT)
    ws = wb["Account API"]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    style_src = 1
    for i, row in enumerate(rows, start=2):
        if i > ws.max_row:
            ws.append([None] * 13)
        for c in range(1, 14):
            src = ws.cell(style_src, c)
            dst = ws.cell(i, c)
            if src.has_style:
                dst.font = copy(src.font)
                dst.border = copy(src.border)
                dst.fill = copy(src.fill)
                dst.number_format = copy(src.number_format)
                dst.protection = copy(src.protection)
                dst.alignment = copy(src.alignment)
        for col, key in zip("DEFGH", "DEFGH"):
            ws[f"{col}{i}"] = row.get(key)
        for col in list("ABC") + list("IJKLM"):
            ws[f"{col}{i}"] = None
    wb.save(OUT)
    return OUT


def main():
    rows_001 = load_001_rows()
    with ROWS_JSON.open(encoding="utf-8") as f:
        rows_rest = json.load(f)
    all_rows = rows_001 + rows_rest
    out = write_xlsx(all_rows)
    counts = {}
    for r in all_rows:
        s = r.get("spec", "?")
        counts[s] = counts.get(s, 0) + 1
    print(f"Wrote {len(all_rows)} rows to {out}")
    for s in sorted(counts):
        print(f"  spec {s}: {counts[s]}")


if __name__ == "__main__":
    main()
