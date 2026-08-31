#!/usr/bin/env python3
"""Generate combined TC xlsx for specs 001, 002, 006, 007, 008, 009, 010 (API focus).

규칙 SSOT: ../SKILL.md
- 중분류(E): API·연동규격 그룹명만. 동일 API = 동일 문자열.
- 시험목적(F): [정상]/[실패]/[변경] + 검증 시나리오.
"""
from copy import copy
from datetime import date
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[4]
TEMPLATE = REPO / ".agents/skills/tc-example-generator/templates/self-test-template.xlsx"
OUTPUT = REPO / f"TC_001-010_api_{date.today().strftime('%Y%m%d')}.xlsx"

# (대분류, 중분류=API그룹, 시험목적, 시험절차, 시험내역)
TC_ROWS = [
    # === 001 ===
    (
        "이메일 인증",
        "6.22 이메일 인증 발송",
        "[정상] flow_token·ID가 유효할 때 resultcode 200과 인증 메일 발송 여부를 확인한다.",
        "1. 회원가입 Redis에 유효한 flow_token을 준비한다.\n2. POST /v1/api/account/email/verifications를 호출한다.\n3. 응답 resultcode를 확인한다.\n4. DB tbl_email_auth_send에 발송 이력이 생겼는지 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/email/verifications\nContent-Type: application/json\n{\n  \"flow_token\": \"{valid_flow_token}\",\n  \"id\": \"newuser@example.com\",\n  \"user_name\": \"홍길동\",\n  \"phone_number\": \"01012345678\",\n  \"ci\": \"{valid_ci}\"\n}\n\n[기대 응답 예시]\nresultcode: 200\n\n[DB 확인 예시]\ntbl_email_auth_send: 해당 이메일(id)로 신규 발송 건 생성",
    ),
    (
        "이메일 인증",
        "6.22 이메일 인증 발송",
        "[실패] 이미 사용 중인 ID로 호출 시 resultcode 701이 나오고 메일이 나가지 않는지 확인한다.",
        "1. DB에 이미 있는 user_id(이메일)를 준비한다.\n2. POST /v1/api/account/email/verifications를 호출한다.\n3. resultcode 701을 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/email/verifications\n{\n  \"flow_token\": \"{valid_flow_token}\",\n  \"id\": \"{duplicate_id}\",\n  \"user_name\": \"홍길동\",\n  \"ci\": \"{valid_ci}\"\n}\n\n[기대 응답 예시]\nresultcode: 701",
    ),
    (
        "이메일 인증",
        "6.22 이메일 인증 발송",
        "[실패] 만료·없는 flow_token으로 호출 시 resultcode 713이 나오는지 확인한다.",
        "1. Redis에서 만료되었거나 없는 flow_token을 준비한다.\n2. POST /v1/api/account/email/verifications를 호출한다.\n3. resultcode 713을 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/email/verifications\n{\n  \"flow_token\": \"{expired_flow_token}\",\n  \"id\": \"newuser@example.com\",\n  \"user_name\": \"홍길동\",\n  \"ci\": \"{valid_ci}\"\n}\n\n[기대 응답 예시]\nresultcode: 713",
    ),
    (
        "이메일 인증",
        "6.23 이메일 인증 확인",
        "[정상] 메일 「인증하기」 payload로 호출 시 resultcode 200과 인증 완료 처리를 확인한다.",
        "1. 6.22로 보낸 메일에서 암호화된 payload를 준비한다.\n2. POST /v1/api/account/email/confirm를 호출한다.\n3. resultcode 200을 확인한다.\n4. tbl_email_auth_send.client_id가 회원가입 Redis의 client_id로 갱신됐는지 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/email/confirm\nContent-Type: application/json\n{\n  \"payload\": \"{encrypted_payload_from_mail}\"\n}\n\n[기대 응답 예시]\nresultcode: 200\n\n[DB 확인 예시]\ntbl_email_auth_send: 해당 발송 건의 client_id 갱신",
    ),
    (
        "이메일 인증",
        "6.23 이메일 인증 확인",
        "[실패] 최신 메일이 아닌 이전 메일 payload로 호출 시 resultcode 736을 확인한다.",
        "1. 같은 이메일로 6.22를 두 번 호출해 최신·이전 발송 건을 만든다.\n2. 이전 메일의 payload로 POST /v1/api/account/email/confirm를 호출한다.\n3. resultcode 736을 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/email/confirm\n{\n  \"payload\": \"{old_mail_payload}\"\n}\n\n[기대 응답 예시]\nresultcode: 736\nmessage: 최종 발송된 이메일이 아닌 메일에서 인증하기 버튼을 눌렀습니다.",
    ),
    (
        "이메일 인증",
        "6.23 이메일 인증 확인",
        "[실패] 잘못된 payload로 호출 시 resultcode 741을 확인한다.",
        "1. 변조되었거나 잘못된 payload를 준비한다.\n2. POST /v1/api/account/email/confirm를 호출한다.\n3. resultcode 741을 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/email/confirm\n{\n  \"payload\": \"{invalid_payload}\"\n}\n\n[기대 응답 예시]\nresultcode: 741",
    ),
    (
        "이메일 인증",
        "6.24 이메일 인증 결과 확인",
        "[정상] 6.23 인증 완료 후 호출 시 resultcode 200을 확인한다.",
        "1. 6.22와 6.23까지 완료된 상태를 준비한다.\n2. POST /v1/api/account/email/confrim/result를 호출한다.\n3. resultcode 200을 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/email/confrim/result\nContent-Type: application/json\n{\n  \"email\": \"newuser@example.com\",\n  \"flow_token\": \"{same_flow_token_as_6_22}\"\n}\n\n[기대 응답 예시]\nresultcode: 200",
    ),
    (
        "이메일 인증",
        "6.24 이메일 인증 결과 확인",
        "[실패] 6.23 인증 전 호출 시 resultcode 747을 확인한다.",
        "1. 6.22 발송만 하고 6.23은 하지 않은 상태를 준비한다.\n2. POST /v1/api/account/email/confrim/result를 호출한다.\n3. resultcode 747을 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/email/confrim/result\n{\n  \"email\": \"newuser@example.com\",\n  \"flow_token\": \"{flow_token}\"\n}\n\n[기대 응답 예시]\nresultcode: 747\nmessage: 이메일 인증이 완료되지 않았습니다.",
    ),
    # === 002 ===
    (
        "회원가입",
        "mailercheck 검증 API (삭제)",
        "[실패] 제거된 mailercheck API 호출 시 HTTP 404를 확인한다.",
        "1. POST /v1/api/account/mailercheck/single을 호출한다.\n2. HTTP 404를 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/mailercheck/single\n\n[기대 응답 예시]\nHTTP 404",
    ),
    (
        "회원가입",
        "ID 중복 확인",
        "[변경] 응답에 mailerCheckStatus 필드가 없고 resultcode 200 또는 701만 나오는지 확인한다.",
        "1. 유효한 flowToken, userName, ci와 이메일 형식 id를 준비한다.\n2. GET /v1/api/account/id를 호출한다.\n3. 응답 JSON에 mailerCheckStatus 키가 없는지 확인한다.\n4. resultcode가 200 또는 701인지 확인한다.",
        "[요청 예시]\nGET /v1/api/account/id?id={email_id}&flowToken={flow_token}&userName={name}&ci={ci}\n\n[기대 응답 예시]\nresultcode: 200 또는 701\nmailerCheckStatus: (필드 없음)",
    ),
    (
        "회원가입",
        "ID 생성·가입",
        "[변경] 가입 응답에 mailerCheckStatus가 없고 resultcode 1006이 나오지 않는지 확인한다.",
        "1. 가입 가능한 신규 사용자 데이터를 준비한다.\n2. POST /v1/api/account/id를 호출한다.\n3. 응답에 mailerCheckStatus 키가 없는지 확인한다.\n4. resultcode가 1006이 아닌지 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/id\n(기존 가입 요청 본문, 산출물 기준 확인 필요)\n\n[기대 응답 예시]\nmailerCheckStatus: (필드 없음)\nresultcode: 1006 미발생",
    ),
    (
        "회원정보 변경",
        "회원정보 변경",
        "[변경] 대표이메일 변경 시 mailerCheck 없이 기존 검증 규칙만 적용되는지 확인한다.",
        "1. 로그인한 사용자로 POST /v1/api/account를 호출한다.\n2. resultcode 1004, 1006이 나오지 않는지 확인한다.",
        "[요청 예시]\nPOST /v1/api/account\n(대표이메일 변경 본문, 산출물 기준 확인 필요)\n\n[기대 응답 예시]\nresultcode: 1004, 1006 미발생",
    ),
    # === 006 ===
    (
        "SNS 연동 해제 알림",
        "8.17 네이버 연동 해제 알림",
        "[정상] 유효한 HMAC·encryptUniqueId로 호출 시 HTTP 204와 tbl_sns_info 행 삭제를 확인한다.",
        "1. sns_type=3으로 연동된 테스트 계정을 준비한다.\n2. POST /v1/account/sns/naver/disconnect-receive를 폼 형식으로 호출한다.\n3. HTTP 204를 확인한다.\n4. tbl_sns_info에서 해당 sns_id 행이 삭제됐는지 확인한다.",
        "[요청 예시]\nPOST /v1/account/sns/naver/disconnect-receive\nContent-Type: application/x-www-form-urlencoded\nclientId={naver_client_id}&encryptUniqueId={encrypted_sns_id}&timestamp={epoch_ms}&signature={hmac_signature}\n\n[기대 응답 예시]\nHTTP 204\n\n[DB 확인 예시]\ntbl_sns_info: sns_type=3, sns_id={decrypted_id} / 행 삭제",
    ),
    (
        "SNS 연동 해제 알림",
        "8.17 네이버 연동 해제 알림",
        "[실패] HMAC 오류 시 HTTP 401과 DB 미변경을 확인한다.",
        "1. sns_type=3 연동 계정을 준비한다.\n2. 잘못된 signature로 POST /v1/account/sns/naver/disconnect-receive를 호출한다.\n3. HTTP 401을 확인한다.\n4. tbl_sns_info 행이 그대로인지 확인한다.",
        "[요청 예시]\nPOST /v1/account/sns/naver/disconnect-receive\nsignature={invalid_hmac}\n\n[기대 응답 예시]\nHTTP 401\n\n[DB 확인 예시]\ntbl_sns_info: 변경 없음",
    ),
    (
        "SNS 연동 해제 알림",
        "8.18 카카오 연동 해제 알림",
        "[정상] 유효한 app_id·user_id로 호출 시 HTTP 200과 행 삭제를 확인한다.",
        "1. sns_type=4 연동 계정을 준비한다.\n2. GET /v1/account/sns/kakao/disconnect-receive?user_id={kakao_user_id}&app_id={kakao_client_id}를 호출한다.\n3. HTTP 200을 확인한다.\n4. tbl_sns_info 행 삭제를 확인한다.",
        "[요청 예시]\nGET /v1/account/sns/kakao/disconnect-receive?user_id={user_id}&app_id={kakao_client_id}\n\n[기대 응답 예시]\nHTTP 200\n\n[DB 확인 예시]\ntbl_sns_info: sns_type=4 / 행 삭제",
    ),
    (
        "SNS 연동 해제 알림",
        "8.18 카카오 연동 해제 알림",
        "[실패] app_id 불일치 시 HTTP 200이지만 DB가 바뀌지 않는지 확인한다.",
        "1. sns_type=4 연동 계정을 준비한다.\n2. 잘못된 app_id로 GET /v1/account/sns/kakao/disconnect-receive를 호출한다.\n3. HTTP 200을 확인한다.\n4. tbl_sns_info 행이 그대로인지 확인한다.",
        "[요청 예시]\nGET /v1/account/sns/kakao/disconnect-receive?user_id={user_id}&app_id={wrong_app_id}\n\n[기대 응답 예시]\nHTTP 200\n\n[DB 확인 예시]\ntbl_sns_info: 변경 없음",
    ),
    (
        "SNS 연동 해제 알림",
        "8.19 애플 연동 해제 알림",
        "[정상] 유효한 JWT payload로 호출 시 HTTP 200과 sns_type=5 행 삭제를 확인한다.",
        "1. sns_type=5 연동 계정을 준비한다.\n2. POST /v1/account/sns/apple/disconnect-receive에 서명된 JWT를 전달한다.\n3. HTTP 200을 확인한다.\n4. tbl_sns_info 행 삭제를 확인한다.",
        "[요청 예시]\nPOST /v1/account/sns/apple/disconnect-receive\nContent-Type: application/json\n{\n  \"payload\": \"{signed_apple_jwt}\"\n}\n\n[기대 응답 예시]\nHTTP 200\n\n[DB 확인 예시]\ntbl_sns_info: sns_type=5, sns_id={sub} / 행 삭제",
    ),
    (
        "SNS 연동 해제 알림",
        "8.19 애플 연동 해제 알림",
        "[실패] JWT aud 오류 시 HTTP 401을 확인한다.",
        "1. aud가 맞지 않는 애플 JWT 테스트 값을 준비한다.\n2. POST /v1/account/sns/apple/disconnect-receive를 호출한다.\n3. HTTP 401을 확인한다.",
        "[요청 예시]\nPOST /v1/account/sns/apple/disconnect-receive\n{\n  \"payload\": \"{jwt_wrong_aud}\"\n}\n\n[기대 응답 예시]\nHTTP 401",
    ),
    (
        "SNS 간편로그인",
        "12.3 SNS 연동 설정",
        "[정상] active_yn=N 호출 시 SNS 제공자 연동 해제 후 tbl_sns_info 행 삭제를 확인한다.",
        "1. SNS가 연동된 계정과 유효한 접속 토큰을 준비한다.\n2. POST /v1/api/account/snsinfo에 active_yn=N을 전달한다.\n3. resultcode 200을 확인한다.\n4. GET /v1/api/account/snsinfo에서 beforeConn=0을 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/snsinfo\nAuthorization: Bearer {access_token}\n{\n  \"sns_type\": 3,\n  \"sns_token\": \"{sns_access_token}\",\n  \"active_yn\": \"N\",\n  \"client_id\": \"{client_id}\"\n}\n\n[기대 응답 예시]\nresultcode: 200\n\n[DB 확인 예시]\ntbl_sns_info: 해당 sns_type 행 삭제",
    ),
    (
        "SNS 간편로그인",
        "12.3 SNS 연동 설정",
        "[실패] SNS 제공자 연동 해제 실패 시 resultcode 720/721/722와 DB 미변경을 확인한다.",
        "1. SNS 제공자 연동 해제가 실패하는 상황을 준비한다.\n2. POST /v1/api/account/snsinfo에 active_yn=N을 전달한다.\n3. resultcode 720, 721, 722 중 하나를 확인한다.\n4. tbl_sns_info 행이 그대로인지 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/snsinfo\nactive_yn=N\n\n[기대 응답 예시]\nresultcode: 720 또는 721 또는 722\n\n[DB 확인 예시]\ntbl_sns_info: 변경 없음",
    ),
    (
        "SNS 간편로그인",
        "6.20 SNS 연동 설정(CI)",
        "[실패] active_yn=N 요청이 HTTP 400으로 거절되는지 확인한다.",
        "1. 유효한 CI와 SNS 토큰을 준비한다.\n2. POST /v1/api/account/id/snsinfo에 active_yn=N을 전달한다.\n3. HTTP 400(유효성 검사 실패)을 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/id/snsinfo\n{\n  \"ci\": \"{valid_ci}\",\n  \"sns_type\": 3,\n  \"active_yn\": \"N\",\n  \"sns_token\": \"{token}\",\n  \"client_id\": \"{client_id}\"\n}\n\n[기대 응답 예시]\nHTTP 400",
    ),
    (
        "회원 탈퇴",
        "6.13 통합 회원 탈퇴",
        "[변경] 탈퇴 시 해당 id_key의 tbl_sns_info가 모두 삭제되는지 확인한다.",
        "1. SNS를 여러 개 연동한 계정을 준비한다.\n2. POST /v1/api/account/withdraw를 호출한다.\n3. 탈퇴가 성공했는지 확인한다.\n4. 해당 id_key의 tbl_sns_info 건수가 0인지 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/withdraw\n(기존 탈퇴 요청 본문, 산출물 기준 확인 필요)\n\n[DB 확인 예시]\ntbl_sns_info: id_key={id_key} / 건수 0",
    ),
    # === 007 ===
    (
        "해외 로그인 차단",
        "12.1 통합 ID 관리 정보 조회",
        "[변경] 응답 login_block_status가 DB overseas_login_block과 같은지 확인한다.",
        "1. overseas_login_block=1인 계정으로 로그인한다.\n2. GET /v1/api/account/userloginmanage/info를 호출한다.\n3. login_block_status=1을 확인한다.",
        "[요청 예시]\nGET /v1/api/account/userloginmanage/info\nAuthorization: Bearer {access_token}\n\n[기대 응답 예시]\n{\n  \"resultcode\": 200,\n  \"message\": \"성공\",\n  \"simple_status\": 1,\n  \"login_block_status\": 1\n}",
    ),
    (
        "해외 로그인 차단",
        "12.4 해외로그인 차단 설정",
        "[정상] login_block_status=1 설정 시 DB와 12.1 조회값이 1인지 확인한다.",
        "1. 유효한 접속 토큰을 준비한다.\n2. POST /v1/api/account/overseas-login-block에 login_block_status=1을 전달한다.\n3. resultcode 200을 확인한다.\n4. tbl_user.overseas_login_block=1을 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/overseas-login-block\nAuthorization: Bearer {access_token}\n{\n  \"login_block_status\": 1\n}\n\n[기대 응답 예시]\nresultcode: 200\n\n[DB 확인 예시]\ntbl_user.overseas_login_block: 1",
    ),
    (
        "해외 로그인 차단",
        "12.4 해외로그인 차단 설정",
        "[실패] 회원 정보 없음 시 resultcode 703을 확인한다.",
        "1. 회원 정보가 없는 접속 토큰 상황을 준비한다.\n2. POST /v1/api/account/overseas-login-block을 호출한다.\n3. resultcode 703을 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/overseas-login-block\n{\"login_block_status\": 1}\n\n[기대 응답 예시]\nresultcode: 703",
    ),
    (
        "해외 로그인 차단",
        "6.25 해외로그인 차단 OTP 알림톡 발송",
        "[정상] 차단 설정 계정으로 호출 시 resultcode 200과 Redis OTP 저장을 확인한다.",
        "1. overseas_login_block=1이고 mobile_no가 있는 계정을 준비한다.\n2. POST /v1/api/account/overseas-login-block/otp/alarm을 호출한다.\n3. resultcode 200을 확인한다.\n4. Redis OTP와 tbl_argos_alarm_history 저장을 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/overseas-login-block/otp/alarm\n{\n  \"id\": \"user_example\"\n}\n\n[기대 응답 예시]\nresultcode: 200",
    ),
    (
        "해외 로그인 차단",
        "6.25 해외로그인 차단 OTP 알림톡 발송",
        "[실패] 차단 미설정 계정으로 호출 시 resultcode 742를 확인한다.",
        "1. overseas_login_block=0인 계정을 준비한다.\n2. POST /v1/api/account/overseas-login-block/otp/alarm을 호출한다.\n3. resultcode 742를 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/overseas-login-block/otp/alarm\n{\"id\": \"{unblocked_user}\"}\n\n[기대 응답 예시]\nresultcode: 742",
    ),
    (
        "해외 로그인 차단",
        "6.25 해외로그인 차단 OTP 알림톡 발송",
        "[실패] mobile_no 없음 시 resultcode 743을 확인한다.",
        "1. overseas_login_block=1이고 mobile_no가 비어 있는 계정을 준비한다.\n2. POST /v1/api/account/overseas-login-block/otp/alarm을 호출한다.\n3. resultcode 743을 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/overseas-login-block/otp/alarm\n{\"id\": \"{no_mobile_user}\"}\n\n[기대 응답 예시]\nresultcode: 743",
    ),
    (
        "해외 로그인 차단",
        "6.26 해외로그인 차단 OTP 검증",
        "[정상] 맞는 OTP로 호출 시 resultcode 200과 overseas_login_block=0을 확인한다.",
        "1. 6.25로 OTP를 발송한다.\n2. POST /v1/api/account/overseas-login-block/otp/check에 id와 otp를 전달한다.\n3. resultcode 200과 message 「OTP 인증 완료」를 확인한다.\n4. tbl_user.overseas_login_block=0과 Redis 삭제를 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/overseas-login-block/otp/check\n{\n  \"id\": \"user_example\",\n  \"otp\": \"123456\"\n}\n\n[기대 응답 예시]\nresultcode: 200\nmessage: OTP 인증 완료\n\n[DB 확인 예시]\ntbl_user.overseas_login_block: 0",
    ),
    (
        "해외 로그인 차단",
        "6.26 해외로그인 차단 OTP 검증",
        "[실패] OTP 불일치 시 resultcode 737과 fail_count를 확인한다.",
        "1. 6.25 발송 후 잘못된 otp로 6.26을 호출한다.\n2. resultcode 737과 fail_count(누적 실패 횟수)를 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/overseas-login-block/otp/check\n{\"id\": \"user_example\", \"otp\": \"000000\"}\n\n[기대 응답 예시]\nresultcode: 737\nmessage: OTP 인증에 실패했습니다.\nfail_count: (1~4, 누적 횟수)",
    ),
    (
        "해외 로그인 차단",
        "6.26 해외로그인 차단 OTP 검증",
        "[실패] OTP 5회 실패 시 resultcode 738을 확인한다.",
        "1. 6.25 발송 후 잘못된 otp를 5번 연속 호출한다.\n2. 5번째 resultcode 738을 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/overseas-login-block/otp/check\n(5회 연속 오입력)\n\n[기대 응답 예시]\nresultcode: 738\nfail_count: 5",
    ),
    (
        "해외 로그인 차단",
        "6.26 해외로그인 차단 OTP 검증",
        "[실패] OTP 입력 시간 만료 시 resultcode 739를 확인한다.",
        "1. 6.25 발송 후 Redis TTL이 만료될 때까지 기다린다.\n2. POST /v1/api/account/overseas-login-block/otp/check를 호출한다.\n3. resultcode 739를 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/overseas-login-block/otp/check\n{\"id\": \"user_example\", \"otp\": \"123456\"}\n\n[기대 응답 예시]\nresultcode: 739",
    ),
    (
        "해외 로그인 차단",
        "11.5 해외로그인 차단 설정(관리자)",
        "[정상] login_block_status 설정 시 DB overseas_login_block 변경을 확인한다.",
        "1. 존재하는 user_id를 준비한다.\n2. POST /v1/api/account/admin/overseas-login-block을 호출한다.\n3. resultcode 200을 확인한다.\n4. tbl_user.overseas_login_block 값을 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/admin/overseas-login-block\n{\n  \"id\": \"user_example\",\n  \"login_block_status\": 1\n}\n\n[기대 응답 예시]\nresultcode: 200\n\n[DB 확인 예시]\ntbl_user.overseas_login_block: 1",
    ),
    (
        "해외 로그인 차단",
        "11.5 해외로그인 차단 설정(관리자)",
        "[실패] 없는 id로 호출 시 resultcode 700을 확인한다.",
        "1. POST /v1/api/account/admin/overseas-login-block에 없는 id를 전달한다.\n2. resultcode 700을 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/admin/overseas-login-block\n{\"id\": \"nonexistent\", \"login_block_status\": 1}\n\n[기대 응답 예시]\nresultcode: 700",
    ),
    # === 008 ===
    (
        "SNS 간편로그인",
        "구글 프로필 조회",
        "[정상] 유효한 id_token으로 호출 시 resultcode 200과 프로필 data를 확인한다.",
        "1. 유효한 구글 id_token과 clientId를 준비한다.\n2. POST /v1/api/account/sns/google/profile를 호출한다.\n3. resultcode 200과 data.sub, email 등을 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/sns/google/profile\nAuthorization: Bearer {access_token}\n{\n  \"clientId\": \"{google_client_id}\",\n  \"idToken\": \"{valid_id_token}\"\n}\n\n[기대 응답 예시]\nresultcode: 200\ndata: sub, email, emailVerified, name, picture, iss, aud, exp",
    ),
    (
        "SNS 간편로그인",
        "구글 프로필 조회",
        "[실패] id_token 오류 시 resultcode 744를 확인한다.",
        "1. 변조된 id_token을 준비한다.\n2. POST /v1/api/account/sns/google/profile를 호출한다.\n3. resultcode 744를 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/sns/google/profile\n{\"clientId\": \"{client_id}\", \"idToken\": \"{invalid_id_token}\"}\n\n[기대 응답 예시]\nresultcode: 744",
    ),
    (
        "SNS 간편로그인",
        "12.3 SNS 연동 설정",
        "[정상] sns_type=6, active_yn=Y 연동 시 tbl_sns_info 저장을 확인한다.",
        "1. 유효한 접속 토큰, id_token, refresh_token을 준비한다.\n2. POST /v1/api/account/snsinfo에 sns_type=6, active_yn=Y를 전달한다.\n3. resultcode 200을 확인한다.\n4. GET /v1/api/account/snsinfo에서 sns_type=6, beforeConn=1을 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/snsinfo\n{\n  \"sns_type\": 6,\n  \"active_yn\": \"Y\",\n  \"sns_token\": \"{id_token}\",\n  \"refresh_token\": \"{refresh_token}\",\n  \"client_id\": \"{client_id}\"\n}\n\n[기대 응답 예시]\nresultcode: 200\n\n[DB 확인 예시]\ntbl_sns_info: sns_type=6 행 저장",
    ),
    (
        "SNS 간편로그인",
        "6.20 SNS 연동 설정(CI)",
        "[정상] sns_type=6, active_yn=Y CI 연동 성공을 확인한다.",
        "1. 유효한 CI와 id_token을 준비한다.\n2. POST /v1/api/account/id/snsinfo를 호출한다.\n3. resultcode 200을 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/id/snsinfo\n{\n  \"ci\": \"{valid_ci}\",\n  \"sns_type\": 6,\n  \"active_yn\": \"Y\",\n  \"sns_token\": \"{id_token}\",\n  \"client_id\": \"{client_id}\"\n}\n\n[기대 응답 예시]\nresultcode: 200",
    ),
    (
        "SNS 간편로그인",
        "12.2 SNS 연동 목록",
        "[변경] list에 sns_type=6 항목이 포함되는지 확인한다.",
        "1. 구글을 연동한 계정으로 GET /v1/api/account/snsinfo를 호출한다.\n2. list에 sns_type=6, beforeConn, activeYn, snsId가 있는지 확인한다.",
        "[요청 예시]\nGET /v1/api/account/snsinfo\nAuthorization: Bearer {access_token}\n\n[기대 응답 예시]\nlist 항목: sns_type=6, beforeConn, activeYn, snsId",
    ),
    # === 009 ===
    (
        "SNS 간편로그인",
        "12.3 SNS 연동 설정",
        "[실패] 미등록 client_id로 구글 Y 연동 시 resultcode 803을 확인한다.",
        "1. 등록되지 않은 client_id와 유효한 id_token을 준비한다.\n2. POST /v1/api/account/snsinfo에 sns_type=6, active_yn=Y를 전달한다.\n3. resultcode 803을 확인한다.\n4. tbl_sns_info에 저장되지 않았는지 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/snsinfo\n{\n  \"sns_type\": 6,\n  \"active_yn\": \"Y\",\n  \"sns_token\": \"{id_token}\",\n  \"client_id\": \"{unregistered_client_id}\",\n  \"refresh_token\": \"{refresh_token}\"\n}\n\n[기대 응답 예시]\nresultcode: 803",
    ),
    (
        "SNS 간편로그인",
        "12.3 SNS 연동 설정",
        "[실패] 구글 Y 연동 시 refresh_token 누락하면 HTTP 400을 확인한다.",
        "1. refresh_token 없이 POST /v1/api/account/snsinfo 구글 Y를 호출한다.\n2. HTTP 400과 유효성 검사 실패 메시지를 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/snsinfo\n(sns_type=6, active_yn=Y, refresh_token 생략)\n\n[기대 응답 예시]\nHTTP 400\nmessage: 유효성 검사 실패 : refresh_token",
    ),
    (
        "SNS 간편로그인",
        "12.3 SNS 연동 설정",
        "[정상] 구글 Y 연동 성공 시 revoke_refresh_token 저장을 확인한다.",
        "1. 유효한 id_token과 refresh_token으로 12.3 구글 Y를 호출한다.\n2. resultcode 200을 확인한다.\n3. tbl_sns_info.revoke_refresh_token 값을 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/snsinfo\n(구글 Y + refresh_token)\n\n[DB 확인 예시]\ntbl_sns_info: revoke_refresh_token={refresh_token}",
    ),
    (
        "SNS 간편로그인",
        "12.3 SNS 연동 설정",
        "[정상] 구글 N 연동 해제 시 revoke_refresh_token으로 행 삭제를 확인한다.",
        "1. revoke_refresh_token이 저장된 구글 연동 계정을 준비한다.\n2. POST /v1/api/account/snsinfo에 sns_type=6, active_yn=N을 전달한다.\n3. resultcode 200과 행 삭제를 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/snsinfo\n{\n  \"sns_type\": 6,\n  \"active_yn\": \"N\",\n  \"sns_token\": \"{id_token}\",\n  \"client_id\": \"{client_id}\"\n}\n\n[기대 응답 예시]\nresultcode: 200\n\n[DB 확인 예시]\ntbl_sns_info: sns_type=6 / 행 삭제",
    ),
    (
        "SNS 간편로그인",
        "12.3 SNS 연동 설정",
        "[실패] revoke_refresh_token 없이 구글 N 호출 시 resultcode 745를 확인한다.",
        "1. revoke_refresh_token이 없는 구글 연동 행을 준비한다.\n2. POST /v1/api/account/snsinfo에 active_yn=N을 전달한다.\n3. resultcode 745를 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/snsinfo\nactive_yn=N, sns_type=6\n\n[기대 응답 예시]\nresultcode: 745\n\n[DB 확인 예시]\ntbl_sns_info: 변경 없음",
    ),
    (
        "SNS 간편로그인",
        "12.3 SNS 연동 설정",
        "[실패] 애플 id_token aud 불일치 시 resultcode 718을 확인한다.",
        "1. aud가 맞지 않는 애플 id_token을 준비한다.\n2. POST /v1/api/account/snsinfo에 sns_type=5, active_yn=Y를 전달한다.\n3. resultcode 718을 확인한다.",
        "[요청 예시]\nPOST /v1/api/account/snsinfo\nsns_type=5, active_yn=Y\n\n[기대 응답 예시]\nresultcode: 718",
    ),
    # === 010 ===
    (
        "구글 RISC 수신",
        "8.20 구글 RISC 연동 해제 알림",
        "[정상] 유효한 보안 이벤트 JWT로 호출 시 HTTP 202와 sns_type=6 행 삭제를 확인한다.",
        "1. sns_type=6 연동 계정과 유효한 RISC JWT 테스트 값을 준비한다.\n2. POST /v1/account/sns/google/disconnect-receive에 JWT 문자열만 body로 전달한다.\n3. HTTP 202를 확인한다.\n4. tbl_sns_info 행 삭제를 확인한다.",
        "[요청 예시]\nPOST /v1/account/sns/google/disconnect-receive\nContent-Type: application/secevent+jwt\n{signed_jwt_string}\n\n[기대 응답 예시]\nHTTP 202\n\n[DB 확인 예시]\ntbl_sns_info: sns_type=6, sns_id={subject.sub} / 행 삭제",
    ),
    (
        "구글 RISC 수신",
        "8.20 구글 RISC 연동 해제 알림",
        "[실패] aud 미등록 JWT로 호출 시 HTTP 401을 확인한다.",
        "1. 등록되지 않은 aud를 가진 RISC JWT 테스트 값을 준비한다.\n2. POST /v1/account/sns/google/disconnect-receive를 호출한다.\n3. HTTP 401을 확인한다.",
        "[요청 예시]\nPOST /v1/account/sns/google/disconnect-receive\n(미등록 aud JWT)\n\n[기대 응답 예시]\nHTTP 401",
    ),
    (
        "구글 RISC 수신",
        "8.20 구글 RISC 연동 해제 알림",
        "[실패] body 비어 있음·JWT 형식 아님 시 HTTP 400을 확인한다.",
        "1. 빈 body 또는 잘못된 문자열로 POST를 호출한다.\n2. HTTP 400을 확인한다.",
        "[요청 예시]\nPOST /v1/account/sns/google/disconnect-receive\n(body: empty)\n\n[기대 응답 예시]\nHTTP 400",
    ),
    (
        "구글 RISC 수신",
        "8.20 구글 RISC 연동 해제 알림",
        "[정상] 처리 대상 아닌 이벤트만 있어도 HTTP 202·DB 유지를 확인한다.",
        "1. verification 등 처리 대상이 아닌 이벤트만 담은 JWT를 준비한다.\n2. POST /v1/account/sns/google/disconnect-receive를 호출한다.\n3. HTTP 202를 확인한다.\n4. tbl_sns_info가 그대로인지 확인한다.",
        "[요청 예시]\nPOST /v1/account/sns/google/disconnect-receive\n(비대상 이벤트 JWT)\n\n[기대 응답 예시]\nHTTP 202\n\n[DB 확인 예시]\ntbl_sns_info: 변경 없음",
    ),
    (
        "구글 RISC 수신",
        "8.20 구글 RISC 연동 해제 알림",
        "[실패] iss 또는 서명 오류 시 HTTP 401을 확인한다.",
        "1. iss 또는 서명이 맞지 않는 JWT를 준비한다.\n2. POST /v1/account/sns/google/disconnect-receive를 호출한다.\n3. HTTP 401을 확인한다.",
        "[요청 예시]\nPOST /v1/account/sns/google/disconnect-receive\n(iss 또는 서명 오류 JWT)\n\n[기대 응답 예시]\nHTTP 401",
    ),
]


def copy_row_style(ws, src_row: int, dst_row: int, max_col: int = 13) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = copy(src.number_format)
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)


def main() -> None:
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb.active

    style_src_row = 2 if ws.max_row >= 2 else 1

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for row_idx, (cat, subcat, purpose, procedure, detail) in enumerate(TC_ROWS, start=2):
        if row_idx > 2:
            ws.insert_rows(row_idx)
        copy_row_style(ws, style_src_row, row_idx)
        ws.cell(row_idx, 4, cat)
        ws.cell(row_idx, 5, subcat)
        ws.cell(row_idx, 6, purpose)
        ws.cell(row_idx, 7, procedure)
        ws.cell(row_idx, 8, detail)
        for col in list(range(1, 4)) + list(range(9, 14)):
            ws.cell(row_idx, col, None)

    wb.save(OUTPUT)
    print(f"Saved {OUTPUT} with {len(TC_ROWS)} rows")


if __name__ == "__main__":
    main()
