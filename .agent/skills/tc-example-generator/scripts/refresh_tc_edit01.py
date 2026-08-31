#!/usr/bin/env python3
"""Refresh non-001 TC rows from edit01: E grouping + strip Speckit-only IDs from D~H."""
import json
import re
import shutil
from copy import copy
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[4]
EDIT01 = Path(
    r"c:\Users\user\OneDrive\바탕 화면\통합ID작업\05. 시험(PL, 개발자, QL, ML)"
    r"\TC_001-010_merged_20260610-edit01.xlsx"
)
TEMPLATE = ROOT / ".agents/skills/tc-example-generator/templates/self-test-template.xlsx"
OUT_JSON = Path(__file__).parent / "tc_rows_002_010_refreshed.json"
OUT_XLSX = ROOT / f"TC_001-010_merged_{date.today():%Y%m%d}.xlsx"
KEEP_D = "회원가입 이메일 인증"

# Speckit 문서 참조·메타 정리 (연동규격 API 번호 6.22·12.3 등은 유지)
_POST_CLEANUPS = [
    (re.compile(r"\bNC에\s*따라"), "제약에 따라"),
    (re.compile(r"\(NC\)"), ""),
    (re.compile(r"smoke-phase\d*\.md", re.I), ""),
    (re.compile(r"smoke-phase\d*", re.I), ""),
    (re.compile(r"plan·"), ""),
    (re.compile(r"plan\s*§[^\s·]+"), ""),
    (re.compile(r"시나리오\s+시나리오"), "시나리오"),
    (re.compile(r"체크리스트\s+체크리스트"), "체크리스트"),
    (re.compile(r"배포 스모크[^\n]*시나리오\s*1\s*~\s*9[^\n]*"), ""),
]

# spec 내부 스모크·plan 참조 행 → 수행자용 E2E 문장으로 교체
_INTEGRATION_ROW = {
    "D": "해외 로그인 차단",
    "E": "해외 로그인 차단 통합 시나리오",
    "F": (
        "12.1·12.4·6.25·6.26·11.5를 순서대로 호출하여 "
        "DB·Redis·응답 코드가 일치하는지 통합 확인한다."
    ),
    "G": (
        "1. 차단 중 이용자로 12.1을 호출해 login_block_status=1을 확인한다.\n"
        "2. 12.4로 차단·해제(0/1) 후 DB·12.1 값이 일치하는지 확인한다(없는 회원 resultcode 703).\n"
        "3. 6.25: 미차단 742·휴대폰 번호 없음 743·정상 발송(Redis·Argos·이력)·동일 id 재발송(OTP·fail_count 리셋)을 확인한다.\n"
        "4. 6.26: 정상 otp 시 resultcode 200·DB overseas_login_block=0, 오입력 737/738, 만료 739를 확인한다.\n"
        "5. 11.5로 id별 차단·해제 및 없는 id resultcode 700을 확인한다."
    ),
    "H": (
        "[기대 응답 예시]\n"
        "상태 조회·설정·OTP 발송·OTP 인증·관리자 설정 API resultcode/HTTP가 계약과 일치\n\n"
        "[DB·Redis 확인 예시]\n"
        "tbl_user.overseas_login_block, tbl_argos_alarm_history, Redis OTP·fail_count"
    ),
}
_STRIP_PATTERNS = [
    re.compile(r"\bFR-\d+(?:-\w+)?\b"),
    re.compile(r"\bSC-\d+\b"),
    re.compile(r"\bR\d+\b"),
    re.compile(r"\bT\d{3}\b"),
    re.compile(r"009 §\d+"),
    re.compile(r"004 brownfield"),
    re.compile(r"\(FR-[^)]+\)"),
    re.compile(r"### 제약 \(NC\)"),
    re.compile(r"\bspecs/\d{3}-[\w-]+"),
    re.compile(r"\n\[근거\][\s\S]*?(?=\n\[|$)"),
    re.compile(r"\n\(근거\)[^\n]*"),
]


def strip_internals(text: str | None) -> str | None:
    if not text or not isinstance(text, str):
        return text
    s = text
    for pat in _STRIP_PATTERNS:
        s = pat.sub("", s)
    for pat, repl in _POST_CLEANUPS:
        s = pat.sub(repl, s)
    s = re.sub(r"  +", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()


def group_e(d: str, old_e: str) -> str:
    e = old_e or ""

    if d == "SNS 연동 해제":
        if e.startswith("네이버"):
            return "네이버 연동 해제 알림 수신"
        if e.startswith("카카오"):
            return "카카오 연동 해제 알림 수신"
        if e.startswith("애플"):
            return "애플 연동 해제 알림 수신"
        if "12.2" in e or "목록" in e:
            return "SNS 연동 목록 조회"
        if "재연동" in e:
            return "SNS 재연동"
        if "위조" in e or "멱등" in e or "공통" in e:
            return "연동 해제 알림 공통"
        return "연동 해제 알림 수신"

    if d == "회원 탈퇴":
        if "관리자" in e or "11.3" in e:
            return "관리자 회원 탈퇴"
        return "회원 탈퇴"

    if d == "SNS 연동 설정":
        if "6.20" in e or "회원가입" in e:
            return "회원가입 SNS 연동"
        if "애플" in e and ("refresh" in e.lower() or "revoke" in e.lower()):
            return "애플 연동 해제"
        if "구글" in e and ("해제" in e or "N" in e or "revoke" in e.lower()):
            return "구글 연동 해제"
        if "구글" in e and ("연결" in e or "Y" in e):
            return "구글 연동 설정"
        if "active_yn=N" in e or "해제" in e or "revoke" in e.lower():
            return "SNS 연동 해제"
        if "Feign" in e or "revoke" in e.lower():
            return "SNS 연동 해제"
        if "미지원" in e or "id_token" in e or "저장" in e:
            return "SNS 연동 설정"
        return "SNS 연동 설정"

    if d == "해외 로그인 차단":
        if "12.1" in e or "상태 조회" in e:
            return "해외 로그인 차단 상태 조회"
        if "12.4" in e or ("설정" in e and "관리자" not in e):
            return "해외 로그인 차단 설정"
        if "11.5" in e or "관리자" in e:
            return "관리자 해외 로그인 차단"
        if "통합" in e or "smoke" in e.lower():
            return "해외 로그인 차단 통합 시나리오"
        return "해외 로그인 차단"

    if d == "해외 로그인 차단 OTP":
        if "6.26" in e or "인증" in e:
            return "OTP 인증"
        return "OTP 알림톡 발송"

    if d == "구글 SNS 연동":
        if "프로필" in e:
            return "구글 프로필 조회"
        if "ActionLog" in e:
            return "구글 ActionLog"
        if "토큰" in e:
            return "구글 연동 검증"
        if "해제" in e or "식별" in e:
            return "구글 연동 해제"
        return "구글 SNS 연동"

    if d == "SNS 연동 목록":
        return "SNS 연동 목록 조회"

    if d == "ActionLog":
        if "6.20" in e or "회원가입" in e:
            return "회원가입 SNS ActionLog"
        if "12.3" in e or "구글" in e:
            return "SNS 연동 ActionLog"
        if "네이버" in e or "카카오" in e or "애플" in e:
            return "SNS ActionLog 회귀"
        return "ActionLog 회귀"

    if d == "SNS aud 검증":
        if "애플" in e:
            return "애플 aud 검증"
        if "네이버" in e or "카카오" in e:
            return "네이버·카카오 aud 미적용"
        if "6.20" in e or "회원가입" in e:
            return "회원가입 SNS aud 검증"
        if "clientId" in e or "validation" in e.lower():
            return "공통 요청 검증"
        if "GoogleAuth" in e or "allowlist" in e:
            return "구글 aud 검증"
        return "구글 aud 검증"

    if d == "구글 refresh_token":
        return "구글 refresh_token 저장"

    if d == "구글 연동 해제":
        return "구글 연동 해제"

    if d == "Google RISC 수신":
        return "Google RISC 알림 수신"

    # fallback: strip Speckit refs from E
    cleaned = strip_internals(e) or e
    return cleaned.strip() or e


def load_edit01_rows():
    wb = load_workbook(EDIT01, read_only=True, data_only=True)
    ws = wb["Account API"]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(r[:8]):
            continue
        rows.append({"D": r[3], "E": r[4], "F": r[5], "G": r[6], "H": r[7]})
    wb.close()
    return rows


def _is_internal_smoke_ref(row: dict) -> bool:
    blob = " ".join(str(row.get(k) or "") for k in "DEFGH")
    return bool(
        re.search(r"smoke-phase", blob, re.I)
        or re.search(r"plan·", blob)
        or re.search(r"시나리오\s*1\s*~\s*9", blob)
        or (
            row.get("D") == "해외 로그인 차단"
            and row.get("E") == "해외 로그인 차단 통합 시나리오"
        )
    )


def refresh_row(row: dict) -> dict:
    if row["D"] == KEEP_D:
        return dict(row)
    if _is_internal_smoke_ref(row):
        return dict(_INTEGRATION_ROW)
    new_e = group_e(row["D"], row["E"] or "")
    return {
        "D": row["D"],
        "E": new_e,
        "F": strip_internals(row["F"]),
        "G": strip_internals(row["G"]),
        "H": strip_internals(row["H"]),
    }


def write_xlsx(rows):
    shutil.copy(TEMPLATE, OUT_XLSX)
    wb = load_workbook(OUT_XLSX)
    ws = wb["Account API"]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    style_src = 1
    for i, row in enumerate(rows, start=2):
        if i > ws.max_row:
            ws.append([None] * 13)
        for c in range(1, 14):
            src = ws.cell(style_src, c)
            dst = ws.cell(i, c)
            if src.has_style:
                dst.font = copy(src.font)
                dst.border = copy(src.border)
                dst.fill = copy(src.fill)
                dst.number_format = copy(src.number_format)
                dst.protection = copy(src.protection)
                dst.alignment = copy(src.alignment)
        for col, key in zip("DEFGH", "DEFGH"):
            ws[f"{col}{i}"] = row.get(key)
        for col in list("ABC") + list("IJKLM"):
            ws[f"{col}{i}"] = None
    wb.save(OUT_XLSX)


def main():
    raw = load_edit01_rows()
    refreshed = [refresh_row(r) for r in raw]
    keep = [r for r in refreshed if r["D"] == KEEP_D]
    rest = [r for r in refreshed if r["D"] != KEEP_D]
    OUT_JSON.write_text(
        json.dumps(rest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    write_xlsx(refreshed)
    from collections import Counter

    print(f"Total: {len(refreshed)} (001 kept: {len(keep)}, refreshed: {len(rest)})")
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_JSON}")
    print("\nE groups (non-001):")
    for d, c in Counter(r["D"] for r in rest).most_common():
        es = Counter(x["E"] for x in rest if x["D"] == d)
        print(f"  {d} ({c})")
        for e, n in es.most_common():
            print(f"    {n}x {e}")


if __name__ == "__main__":
    main()
