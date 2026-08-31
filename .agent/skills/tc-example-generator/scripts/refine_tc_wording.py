#!/usr/bin/env python3
"""Refine E(중분류)·F(시험목적) to human-readable style (001 reference)."""
import json
import shutil
from copy import copy
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[4]
ROWS_JSON = Path(__file__).parent / "tc_rows_002_010.json"
USER_001 = Path(r"c:\Users\user\Downloads\TC_001_m365-email-send_20260610.xlsx")
OUT_001 = ROOT / "specs/001-m365-email-send/TC_001_m365-email-send_20260610.xlsx"
TEMPLATE = ROOT / ".agents/skills/tc-example-generator/templates/self-test-template.xlsx"
OUT_MERGED = ROOT / f"TC_001-010_merged_{date.today().strftime('%Y%m%d')}.xlsx"

# (E, F) per spec — order must match tc_rows_002_010.json
REFINED: dict[str, list[tuple[str, str]]] = {
    "002": [
        ("mailercheck 설정 없이 기동 성공", "mailercheck 연동 설정 없이 local 프로필에서 애플리케이션이 정상 기동되는지 확인한다."),
        ("mailercheck single API 제거 확인 (404)", "삭제된 POST /mailercheck/single 호출 시 HTTP 404를 확인한다."),
        ("mailercheck 비동기 발송 API 제거 확인 (404)", "삭제된 POST /mailercheck/single/async 호출 시 HTTP 404를 확인한다."),
        ("mailercheck 비동기 결과 조회 API 제거 확인 (404)", "삭제된 GET /mailercheck/single/async/{id} 호출 시 HTTP 404를 확인한다."),
        ("GET /id 응답 mailerCheckStatus 필드 제거", "이메일 형식 ID 중복 확인 시 mailerCheckStatus 없이 resultcode 200 또는 701만 반환되는지 확인한다."),
        ("GET /id 호출 시 외부 mailerCheck 미사용", "이메일 형식 ID 중복 확인 시 외부 mailerCheck API가 호출되지 않음을 확인한다."),
        ("POST /id 가입 시 mailerCheck 단계 없음", "가입 처리 시 mailerCheckStatus 필드·UTIL_1006이 발생하지 않음을 확인한다."),
        ("POST /api/account 이메일 변경 시 mailerCheck 미사용", "대표이메일 변경 시 UTIL_1004·UTIL_1006이 발생하지 않음을 확인한다."),
        ("mailerCheck 전용 오류코드(1003~1008) 미사용", "계정 API 응답에 UTIL_1003~1008이 반환되지 않음을 확인한다."),
        ("MailerCheckException 응답 형식 미노출", "mailer_check_status 포함 예외 응답이 더 이상 노출되지 않음을 확인한다."),
        ("소스·설정 mailerCheck 잔재 제거 확인", "Util·Feign·AOP 등 NC1~4 항목 제거가 완료되었는지 확인한다."),
        ("GET /id — UTIL_1004 미발생", "이메일 형식 ID 중복 확인 시 UTIL_1004가 반환되지 않음을 확인한다."),
        ("Redis mailerCheck 캐시 신규 적재 중단", "mailerCheck 결과 Redis 신규 적재가 중단되었는지 확인한다."),
        ("mailerCheck 제거 완료 판정(SC-003)", "연동 전용 Util·서비스·Feign·오류코드 제거가 완료되었는지 확인한다."),
    ],
    "006": [
        ("네이버 연동 해제 알림 수신 성공", "유효한 HMAC·복호화 후 sns_type=3 행이 삭제되고 HTTP 204로 응답하는지 확인한다."),
        ("네이버 HMAC 불일치 시 인증 실패", "signature 불일치 시 tbl_sns_info 변경 없이 HTTP 401로 실패를 확인한다."),
        ("네이버 필수 파라미터 누락 시 실패", "clientId·encryptUniqueId·timestamp·signature 누락·빈 값 시 HTTP 400으로 실패를 확인한다."),
        ("네이버 clientId 불일치 시 인증 실패", "설정과 다른 clientId 시 HTTP 401·DB 미변경으로 실패를 확인한다."),
        ("네이버 DB 장애 시 서버 오류", "DB 연결·조회 실패 등 인프라 장애 시 HTTP 500을 확인한다."),
        ("네이버 매칭 행 없음 시 no-op 성공", "검증 성공·매칭 행 없을 때 삭제 없이 HTTP 204로 응답하는지 확인한다."),
        ("카카오 연동 해제 웹훅 수신 성공", "app_id 일치·user_id 매칭 시 sns_type=4 행 삭제 후 HTTP 200으로 응답하는지 확인한다."),
        ("카카오 app_id 불일치 시 DB 미변경", "app_id 불일치 시에도 HTTP 200·tbl_sns_info 미변경을 확인한다."),
        ("카카오 내부 오류 시에도 HTTP 200 응답", "이용자 없음·내부 오류 포함 처리 후 3초 이내 HTTP 200을 확인한다."),
        ("애플 consent-revoked 알림 수신 성공", "payload 검증 성공·events.type=consent-revoked 시 sns_type=5 행 삭제·HTTP 200을 확인한다."),
        ("애플 account-delete 알림 수신 성공", "events.type=account-delete 시 sns_type=5 행 삭제·HTTP 200을 확인한다."),
        ("애플 payload 누락 시 실패", "body에 payload 없음 시 HTTP 400으로 실패를 확인한다."),
        ("애플 iss·aud 불일치 시 인증 실패", "iss·aud 불일치 시 HTTP 401·DB 미변경으로 실패를 확인한다."),
        ("애플 DB 장애 시 서버 오류", "DB 연결·조회 실패 등 인프라 장애 시 HTTP 500을 확인한다."),
        ("12.2 목록 삭제 후 미연결 표시", "인바운드 삭제 후 12.2에서 해당 sns_type이 beforeConn=0·빈 snsId로 표시되는지 확인한다."),
        ("연동 해제 후 동일 SNS 재연동 성공", "행 삭제 후 12.3 active_yn=Y로 재연동이 가능한지 확인한다."),
        ("6.13 통합 회원 탈퇴 시 SNS 선행 삭제", "tbl_user 삭제 전 id_key의 모든 tbl_sns_info 행이 삭제되는지 확인한다."),
        ("11.3 관리자 탈퇴 시 SNS 선행 삭제", "관리자 탈퇴 시에도 deleteAllSnsInfoByIdKey 후 사용자 삭제 순서를 확인한다."),
        ("12.3 active_yn=N 연동 해제 성공", "SNS 제공자 연동 해제 API 성공 시 행 삭제·active_yn=N 갱신 없음을 확인한다."),
        ("12.3 active_yn=N revoke 실패", "SNS 제공자 연동 해제 실패 시 tbl_sns_info 미변경·resultcode 720/721/722로 실패를 확인한다."),
        ("12.3 revoke 성공·DB DELETE 실패", "revoke 성공·DB DELETE 실패 시 HTTP 500·행 유지를 확인한다."),
        ("6.20 active_yn=N 요청 실패", "6.20은 연동(Y) 전용이며 active_yn=N 요청 시 HTTP 400으로 실패를 확인한다."),
        ("위조 알림 수신 시 DB 미변경", "검증 실패 알림 수신 시 tbl_sns_info가 변경되지 않음을 확인한다."),
        ("중복 알림 멱등 처리", "이미 삭제된 행에 재알림 시 no-op·제공자별 응답 정책 유지를 확인한다."),
        ("12.3 애플 N 시 refresh_token으로 revoke", "sns_type=5·active_yn=N 시 refresh_token으로 revoke하는지 확인한다."),
    ],
    "007": [
        ("12.1 해외 로그인 차단 상태 조회 성공(차단 중)", "DB overseas_login_block=1일 때 12.1 login_block_status=1을 확인한다."),
        ("12.1 해외 로그인 차단 상태 조회 성공(해제)", "DB overseas_login_block=0 또는 NULL일 때 login_block_status=0을 확인한다."),
        ("12.1 AccessToken 없음·무효 시 인증 실패", "AccessToken 없음·무효 시 12.1 인증 실패를 확인한다."),
        ("12.4 해외 로그인 차단 설정 성공", "12.4로 login_block_status=1 설정 시 DB·12.1이 1로 일치하는지 확인한다."),
        ("12.4 해외 로그인 차단 해제 성공", "12.4로 login_block_status=0 설정 시 DB·12.1이 0으로 일치하는지 확인한다."),
        ("12.4 동일 값 재설정 성공(멱등)", "이미 차단(1) 상태에서 다시 1 설정 시 성공(멱등)을 확인한다."),
        ("12.4 AccessToken 없음 시 인증 실패", "AccessToken 없이 12.4 호출 시 인증 실패를 확인한다."),
        ("12.4 login_block_status 유효값 외 실패", "login_block_status가 0·1이 아니면 resultcode 400으로 실패를 확인한다."),
        ("12.4 회원 정보 없음 실패", "토큰은 유효하나 회원 없음 시 resultcode 703으로 실패를 확인한다."),
        ("11.5 관리자 차단 설정 성공", "11.5로 특정 user_id 차단(login_block_status=1) 시 DB=1을 확인한다."),
        ("11.5 관리자 차단 해제 성공", "11.5로 OTP 없이 login_block_status=0 해제를 확인한다."),
        ("11.5 존재하지 않는 id 실패", "없는 user_id 시 resultcode 700으로 실패를 확인한다."),
        ("6.25 OTP 알림톡 발송 성공", "차단 중 이용자 6.25 성공 시 Redis·Argos·tbl_argos_alarm_history 저장을 확인한다."),
        ("6.25 5분 내 OTP 재발급 성공", "동일 id 재요청 시 OTP·fail_count=0·TTL 갱신·Argos 재호출을 확인한다."),
        ("6.25 차단 미설정 시 발송 실패", "overseas_login_block=0/NULL 시 resultcode 742·Argos·Redis 미호출로 실패를 확인한다."),
        ("6.25 mobile_no 없음 시 발송 실패", "mobile_no 없음·빈 값 시 resultcode 743·Argos 미호출로 실패를 확인한다."),
        ("6.25 Argos 발송 실패", "Argos 발송 실패 시 resultcode 829·성공 이력 없음을 확인한다."),
        ("6.25 존재하지 않는 id 실패", "없는 id 시 resultcode 700으로 실패를 확인한다."),
        ("6.25 이력 저장 실패", "Argos 성공·이력 INSERT 실패 시 resultcode 829로 실패를 확인한다."),
        ("6.26 OTP 인증 성공·차단 해제", "올바른 otp 입력 시 resultcode 200·OTP 인증 완료·overseas_login_block=0을 확인한다."),
        ("6.26 OTP 불일치 실패(1~4회)", "잘못된 otp 1~4회 누적 시 resultcode 737·fail_count를 확인한다."),
        ("6.26 OTP 5회 초과 실패", "잘못된 otp 5회 이상 시 resultcode 738·fail_count를 확인한다."),
        ("6.26 OTP 만료 실패", "Redis OTP 만료 후 resultcode 739로 실패를 확인한다."),
        ("6.26 존재하지 않는 id 실패", "없는 id 시 resultcode 700으로 실패를 확인한다."),
        ("해외 로그인 차단 통합 시나리오 확인", "plan·smoke-phase7 시나리오 1~9 통합 확인을 수행한다."),
    ],
    "008": [
        ("구글 프로필 조회 성공", "유효 id_token·clientId로 프로필(sub, email 등)이 반환되는지 확인한다."),
        ("구글 프로필 id_token 검증 실패", "id_token 검증 실패 시 resultcode 744로 실패를 확인한다."),
        ("구글 프로필 clientId 오류 실패", "clientId validation·설정 오류 시 resultcode 750 또는 803으로 실패를 확인한다."),
        ("12.3 구글 연결(Y) 성공", "sns_type=6·active_yn=Y·id_token·refreshToken으로 연결 저장·revoke_refresh_token 저장을 확인한다."),
        ("12.3 구글 해제(N) 성공", "sns_token=id_token·DB revoke_refresh_token으로 revoke 성공 시 sns_type=6 행 삭제를 확인한다."),
        ("12.3 구글 해제 revoke 실패", "revoke 실패 시 resultcode 745·행 유지로 실패를 확인한다."),
        ("6.20 구글 연결(Y) 성공", "6.20에서 sns_type=6·active_yn=Y·id_token·ci로 연결 저장을 확인한다."),
        ("6.20 active_yn=N 요청 실패", "6.20은 연동(Y) 전용이며 active_yn=N 요청 시 HTTP 400으로 실패를 확인한다."),
        ("12.2 구글 연결됨 표시", "구글 연결 이용자 12.2에서 sns_type=6 연결 상태를 확인한다."),
        ("12.2 구글 미연결 표시", "해제 후 sns_type=6 미연결·이전 식별 정보 미노출을 확인한다."),
        ("12.2 응답에 sns_account_id 미포함", "12.2·프로필 API 응답에 sns_account_id·연동 일시가 없음을 확인한다."),
        ("12.3 구글 연결·해제 ActionLog 기록", "12.3 구글 처리 시 tbl_user_action_log에 sns_type=6·log_type=407 기록을 확인한다."),
        ("6.20 구글 연결 ActionLog 기록", "6.20 구글 Y 처리 시 ActionLog sns_type=6 기록을 확인한다."),
        ("네이버·카카오·애플 ActionLog 미생성 유지", "sns_type 3·4·5 연결 설정 시 ActionLog 미생성 회귀 없음을 확인한다."),
        ("기존 ActionLog API 회귀 없음", "회원가입·탈퇴 등 기존 ActionLog가 변경·누락되지 않음을 확인한다."),
        ("12.3 미지원 sns_type 실패", "sns_type 3~6 외 요청 시 resultcode 717로 실패를 확인한다."),
        ("12.3 구글 연결 id_token 검증 실패", "연결(Y) 시 id_token 검증 실패 resultcode 744·저장 없음을 확인한다."),
        ("12.3 구글 연결 저장 실패", "중복·insert 실패 시 resultcode 724 또는 729로 실패를 확인한다."),
        ("구글 revoke Feign 호출 확인", "해제(N) 시 GoogleOAuthClient POST oauth2.googleapis.com/revoke 호출을 확인한다."),
        ("토큰 오류 시 DB 미변경", "토큰 검증 실패 시 DB 저장·삭제 없음을 확인한다."),
        ("구글 해제 후 식별 정보 삭제", "해제 성공 후 sns_account_id 등 식별 정보가 남지 않음을 확인한다."),
        ("ActionLog 실패 시 본 처리 유지", "ActionLog 기록 실패가 연동 저장/삭제를 롤백하지 않음을 확인한다."),
    ],
    "009": [
        ("구글 프로필 aud 정합 성공", "tbl_sns_config client_id와 id_token aud 일치 시 프로필 조회 성공을 확인한다."),
        ("구글 aud 불일치 실패", "id_token aud≠tbl_sns_config.client_id 시 resultcode 744로 실패를 확인한다."),
        ("구글 tbl_sns_config 미등록 실패", "sns_type=6 설정 행 없음 시 resultcode 803으로 실패를 확인한다."),
        ("구글 aud 검증 순서 확인", "DB 조회(803) 선행 → GoogleIdTokenVerifier(744) 순서를 확인한다."),
        ("애플 플랫폼 aud 정합 성공", "OAuth app_type·tbl_sns_config ios 행과 aud 일치 시 성공을 확인한다."),
        ("애플 id_token 파싱 실패", "애플 id_token JSON 파싱 실패 시 resultcode 719로 실패를 확인한다."),
        ("애플 aud 불일치 실패", "aud≠확정 client_id 시 resultcode 718로 실패를 확인한다."),
        ("애플 app_type=all 폴백 성공", "플랫폼 행 없음·all 행 존재 시 all 폴백으로 성공을 확인한다."),
        ("애플 설정 없음 실패", "플랫폼·all 행 모두 없음 시 resultcode 803으로 실패를 확인한다."),
        ("애플 client_id 중복 설정 실패", "동일 app_type에 client_id 2건 이상 시 resultcode 746으로 실패를 확인한다."),
        ("clientId validation 실패", "요청 clientId 형식·필수값 실패 시 resultcode 750으로 실패를 확인한다."),
        ("12.3 구글 Y refreshToken 저장 성공", "active_yn=Y·sns_type=6 시 refreshToken 필수·revoke_refresh_token 저장을 확인한다."),
        ("12.3 구글 Y refreshToken 누락 실패", "refreshToken 누락·blank 시 HTTP 400 validation 실패를 확인한다."),
        ("12.3 구글 N 조기 disconnect 성공", "구글 N은 aud·프로필 없이 조기 disconnect하는지 확인한다."),
        ("12.3 구글 N DB revoke_refresh_token 사용", "해제 N 시 DB revoke_refresh_token만으로 Google revoke 후 DELETE를 확인한다."),
        ("12.3 구글 N revoke_refresh_token 없음 실패", "DB revoke_refresh_token null/blank 시 resultcode 745·행 유지로 실패를 확인한다."),
        ("12.3 구글 N revoke API 실패", "Google revoke API 실패 시 resultcode 745·행 유지로 실패를 확인한다."),
        ("네이버·카카오 aud 검증 미적용", "sns_type 3·4 프로필 경로에 aud 로직이 없음을 확인한다."),
        ("6.20 구글·애플 Y aud 검증 적용", "6.20 active_yn=Y 시 구글·애플 aud 검증 경로를 확인한다."),
        ("애플 revoke client_id resolver 적용", "애플 revoke·client_secret JWT sub에 확정 client_id 사용을 확인한다."),
        ("aud 실패 시 후속 처리 없음", "aud 검증 실패 시 저장·revoke·ActionLog 후속 없음을 확인한다."),
        ("GoogleAuthConfig allowlist 폐지", "GOOGLE_CLIENT_ID allowlist 미사용·tbl_sns_config SSOT를 확인한다."),
        ("6.20 구글 refreshToken 미적용", "6.20 구글 Y에 refreshToken 필수·DB 저장 규칙이 적용되지 않음을 확인한다."),
        ("구글 다중 audience aud 정합 성공", "id_token aud 배열에 tbl_sns_config.client_id 포함 시 통과를 확인한다."),
    ],
    "010": [
        ("RISC tokens-revoked 수신 성공", "tokens-revoked·subject.sub 매칭 시 sns_type=6 행 삭제·HTTP 202를 확인한다."),
        ("RISC token-revoked 수신 성공", "token-revoked 이벤트도 동일 삭제·HTTP 202를 확인한다."),
        ("RISC 매칭 행 없음 no-op 성공", "매칭 행 없을 때 DB 미변경·HTTP 202를 확인한다."),
        ("RISC subject.sub 없음 실패", "처리 대상 이벤트에 sub 없음 시 HTTP 400·DB 미변경으로 실패를 확인한다."),
        ("RISC aud 불일치 인증 실패", "aud≠tbl_sns_config client_id 시 HTTP 401으로 실패를 확인한다."),
        ("RISC JWT 만료·서명 실패", "JWT exp 만료 또는 서명 검증 실패 시 HTTP 401·DB 미변경으로 실패를 확인한다."),
        ("RISC iss 불일치 인증 실패", "iss가 https://accounts.google.com(/) 외 시 HTTP 401으로 실패를 확인한다."),
        ("RISC body 형식 오류 실패", "body 없음·blank·비 JWT 형식 시 HTTP 400으로 실패를 확인한다."),
        ("RISC DB 장애 시 서버 오류", "DB 연결·조회 실패 등 인프라 장애 시 HTTP 500을 확인한다."),
        ("RISC 삭제 후 12.2 구글 미연결", "RISC 삭제 후 12.2에서 sns_type=6 beforeConn=0·빈 snsId를 확인한다."),
        ("RISC 동일 알림 재수신 멱등", "이미 삭제된 후 재수신 시 no-op·HTTP 202를 확인한다."),
        ("RISC 비대상 이벤트만 수신 성공", "verification 등 비대상 이벤트만 있으면 DELETE 없이 HTTP 202를 확인한다."),
        ("RISC mixed events 처리 대상만 삭제", "처리 대상·비대상 이벤트 혼합 시 대상 sub만 DELETE·HTTP 202를 확인한다."),
    ],
}


def apply_refined(rows: list[dict]) -> list[dict]:
    by_spec: dict[str, int] = {}
    for row in rows:
        spec = row["spec"]
        idx = by_spec.get(spec, 0)
        pairs = REFINED.get(spec, [])
        if idx < len(pairs):
            row["E"], row["F"] = pairs[idx]
        by_spec[spec] = idx + 1
    return rows


def sync_user_001():
    if USER_001.exists():
        shutil.copy(USER_001, OUT_001)


def write_merged(rows_001, rows_rest):
    all_rows = rows_001 + rows_rest
    shutil.copy(TEMPLATE, OUT_MERGED)
    wb = load_workbook(OUT_MERGED)
    ws = wb["Account API"]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for i, row in enumerate(all_rows, start=2):
        if i > ws.max_row:
            ws.append([None] * 13)
        for col, key in zip("DEFGH", "DEFGH"):
            ws[f"{col}{i}"] = row.get(key)
        for col in list("ABC") + list("IJKLM"):
            ws[f"{col}{i}"] = None
    wb.save(OUT_MERGED)


def load_001_rows():
    path = OUT_001 if OUT_001.exists() else USER_001
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True)
    ws = wb["Account API"]
    rows = []
    for r in range(2, ws.max_row + 1):
        d = ws.cell(r, 4).value
        if not d:
            continue
        rows.append(
            {
                "spec": "001",
                "D": d,
                "E": ws.cell(r, 5).value,
                "F": ws.cell(r, 6).value,
                "G": ws.cell(r, 7).value,
                "H": ws.cell(r, 8).value,
            }
        )
    wb.close()
    return rows


def main():
    sync_user_001()
    with ROWS_JSON.open(encoding="utf-8") as f:
        rows = json.load(f)
    # validate counts
    for spec, pairs in REFINED.items():
        count = sum(1 for r in rows if r["spec"] == spec)
        if count != len(pairs):
            raise SystemExit(f"Count mismatch spec {spec}: json={count} refined={len(pairs)}")
    rows = apply_refined(rows)
    with ROWS_JSON.open("w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)
    rows_001 = load_001_rows()
    write_merged(rows_001, rows)
    print(f"Refined {len(rows)} rows; merged total {len(rows_001)+len(rows)} -> {OUT_MERGED}")


if __name__ == "__main__":
    main()
